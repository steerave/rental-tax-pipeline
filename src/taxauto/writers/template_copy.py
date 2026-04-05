"""Copy a filled XLSX as a template for a new year.

Clears value cells in the specified columns while preserving:
- Labels (column B, C)
- Formulas (e.g., =SUM(D5:D9) in Total Revenues)
- Styling, formatting, sheet structure
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Iterable

import openpyxl


def copy_and_clear_values(
    src: Path,
    dst: Path,
    *,
    value_columns: Iterable[str] = ("D",),
    label_columns: Iterable[str] = ("B", "C"),
) -> Path:
    """Copy an XLSX and clear every non-formula cell in value_columns."""
    src = Path(src)
    dst = Path(dst)
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(src, dst)

    wb = openpyxl.load_workbook(dst)
    for ws in wb.worksheets:
        for col in value_columns:
            for row in range(1, ws.max_row + 1):
                cell = ws[f"{col}{row}"]
                if cell.value is None:
                    continue
                # Preserve formulas — they'll recalculate when opened
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                cell.value = None
    wb.save(dst)
    return dst
