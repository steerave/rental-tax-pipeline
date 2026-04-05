"""Per-property transactions audit tab.

Appends a '{property}_txns' sheet to an output workbook showing every
transaction that contributed to the P&L cell values.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

import openpyxl


HEADERS = ["date", "source", "description", "amount", "template_category", "notes"]


def append_transactions_tab(
    workbook_path: Path,
    *,
    sheet_name: str,
    transactions: Iterable[dict],
) -> None:
    """Add a '{sheet_name}_txns' tab with detailed transaction rows."""
    wb = openpyxl.load_workbook(workbook_path)
    tab_name = f"{sheet_name}_txns"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)
    ws.append(HEADERS)
    for t in transactions:
        ws.append([
            t.get("date"),
            t.get("source", ""),
            t.get("description", ""),
            float(t["amount"]) if t.get("amount") is not None else None,
            t.get("template_category", ""),
            t.get("notes", ""),
        ])
    wb.save(workbook_path)
