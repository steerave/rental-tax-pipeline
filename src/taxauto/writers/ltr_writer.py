"""LTR Excel writer — fills the accountant's per-property P&L template.

Phase 2 design: LTR sheets have drifting row positions between properties.
The writer scans column C per sheet to build a dynamic {label: row} map.
"""

from __future__ import annotations

import warnings
from decimal import Decimal
from pathlib import Path
from typing import Dict, Optional, Sequence

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .template_copy import copy_and_clear_values


def scan_label_rows(ws: Worksheet) -> Dict[str, int]:
    """Scan column C and return {label_text: row_number} for all non-empty cells."""
    out: Dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        v = ws[f"C{row}"].value
        if v is not None:
            label = str(v).strip()
            if label:
                out[label] = row
    return out


def _find_row_case_insensitive(
    label_rows: Dict[str, int], target: str
) -> Optional[int]:
    target_lower = target.lower()
    for label, row in label_rows.items():
        if label.lower() == target_lower:
            return row
    return None


def write_ltr_workbook(
    *,
    template_path: Path,
    output_path: Path,
    per_property_totals: Dict[str, Dict[str, Decimal]],
    year: int,
    property_sheet_names: Sequence[str] = (
        "1015 39th St",
        "1210 College Ave",
        "308 Lincoln Ave",
    ),
) -> Path:
    """Copy the template and fill per-property category totals using dynamic row scan."""
    copy_and_clear_values(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    for sheet_name in property_sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws["B1"] = sheet_name
        ws["D4"] = year

        label_rows = scan_label_rows(ws)
        totals = per_property_totals.get(sheet_name, {})
        for category, value in totals.items():
            if value is None or value == Decimal("0"):
                continue
            row = _find_row_case_insensitive(label_rows, category)
            if row is None:
                warnings.warn(
                    f"LTR writer: category '{category}' not found on sheet '{sheet_name}'"
                )
                continue
            ws[f"D{row}"] = float(value)

    wb.save(output_path)
    return output_path
