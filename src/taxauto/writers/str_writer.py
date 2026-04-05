"""STR Excel writer — fills the accountant's per-property P&L template.

Phase 2 design: all 4 STR sheets share a fixed row layout.
Row positions are hardcoded per the 2024 template discovery.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Dict, Sequence

import openpyxl

from .template_copy import copy_and_clear_values


# Revenue rows (column D)
STR_REVENUE_ROWS = {
    "Sales revenue": 5,
    "(Less sales returns and allowances)": 6,
    "Laundry revenue": 7,
    "late fees revenue": 8,
    "Other revenue": 9,
}

# Expense rows (column D)
STR_EXPENSE_ROWS = {
    "Advertising": 13,
    "Appliances": 14,
    "Bank Charges": 15,
    "Cleaning Fees": 16,
    "Commissions/Service Fees": 17,
    "Furniture and equipment": 18,
    "Insurance": 19,
    "Interest expense": 20,
    "Landscaping": 21,
    "Licenses and Fees": 22,
    "Management Fees": 23,
    "Pest Control": 24,
    "Rent Expense": 25,
    "Renovations": 26,
    "Repairs and Maintenance": 27,
    "legal expenses": 28,
    "Supplies": 29,
    "Travel": 30,
    "Utilities": 31,
    "HOA": 32,
    "other": 33,
}

ALL_STR_ROWS = {**STR_REVENUE_ROWS, **STR_EXPENSE_ROWS}


def write_str_workbook(
    *,
    template_path: Path,
    output_path: Path,
    per_property_totals: Dict[str, Dict[str, Decimal]],
    year: int,
    property_sheet_names: Sequence[str] = (
        "15 Belden",
        "27 Farmstead Dr",
        "20 Valleywood Ln",
        "17 Oak Glen",
    ),
) -> Path:
    """Copy the template and fill per-property category totals."""
    copy_and_clear_values(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    for sheet_name in property_sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws["B1"] = sheet_name
        ws["D4"] = year
        ws["C33"] = "other"  # fix template quirk

        totals = per_property_totals.get(sheet_name, {})
        for category, row in ALL_STR_ROWS.items():
            value = totals.get(category)
            if value is not None and value != Decimal("0"):
                ws[f"D{row}"] = float(value)

    wb.save(output_path)
    return output_path
