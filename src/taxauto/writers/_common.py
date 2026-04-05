"""Helpers shared by the STR and LTR Excel writers."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any, List, Sequence

import openpyxl


COLUMNS = ("Date", "Description", "Amount", "Category")


def copy_template(template_path: Path, output_path: Path) -> Path:
    """Copy the blank template to the output location without modifying it."""
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)
    return output_path


def append_rows(
    output_path: Path,
    rows: Sequence[Sequence[Any]],
    *,
    sheet_name: str | None = None,
) -> None:
    """Append rows to the given sheet (first sheet by default) and save."""
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    for row in rows:
        ws.append(list(row))
    wb.save(output_path)


def locate_header_columns(output_path: Path, *, sheet_name: str | None = None) -> List[str]:
    """Return the header row of the target sheet for sanity checks."""
    wb = openpyxl.load_workbook(output_path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    first_row = next(ws.iter_rows(values_only=True), None) or ()
    return [str(c) if c is not None else "" for c in first_row]
