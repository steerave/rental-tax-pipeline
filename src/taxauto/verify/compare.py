"""Cell-by-cell comparison of filed vs generated workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openpyxl


@dataclass
class ComparisonResult:
    sheet: str
    row: int
    label: str
    filed: Optional[float]
    generated: Optional[float]
    delta: float
    within_tolerance: bool


def compare_workbooks(
    filed_path: Path,
    generated_path: Path,
    *,
    value_column: str = "D",
    label_column: str = "C",
    revenue_tolerance: float = 1.00,
    expense_tolerance_abs: float = 5.00,
    expense_tolerance_pct: float = 0.02,
) -> List[ComparisonResult]:
    """Compare every value cell across all shared sheets."""
    filed_wb = openpyxl.load_workbook(filed_path, data_only=True)
    gen_wb = openpyxl.load_workbook(generated_path, data_only=True)
    results: List[ComparisonResult] = []

    for sheet_name in filed_wb.sheetnames:
        filed_ws = filed_wb[sheet_name]
        gen_ws = gen_wb[sheet_name] if sheet_name in gen_wb.sheetnames else None

        for row in range(1, filed_ws.max_row + 1):
            filed_val = filed_ws[f"{value_column}{row}"].value
            if filed_val is None:
                continue
            if isinstance(filed_val, str):
                if filed_val.startswith("=") or not any(c.isdigit() for c in filed_val):
                    continue
                try:
                    filed_val = float(filed_val)
                except ValueError:
                    continue

            label_cell = filed_ws[f"{label_column}{row}"].value
            label = str(label_cell).strip() if label_cell else f"Row {row}"

            gen_val = 0.0
            if gen_ws is not None:
                gv = gen_ws[f"{value_column}{row}"].value
                if gv is not None and not (isinstance(gv, str) and gv.startswith("=")):
                    try:
                        gen_val = float(gv)
                    except (ValueError, TypeError):
                        gen_val = 0.0

            delta = gen_val - float(filed_val)
            if row <= 10:
                tol = revenue_tolerance
            else:
                tol = max(expense_tolerance_abs, abs(float(filed_val)) * expense_tolerance_pct)

            results.append(ComparisonResult(
                sheet=sheet_name, row=row, label=label,
                filed=float(filed_val), generated=gen_val,
                delta=delta, within_tolerance=abs(delta) <= tol,
            ))

    return results


def format_comparison_table(results: List[ComparisonResult]) -> str:
    """Format comparison results as a markdown table."""
    lines = [
        "| Sheet | Row | Label | Filed | Generated | Delta | OK? |",
        "|---|---|---|---|---|---|---|",
    ]
    for r in results:
        ok = "YES" if r.within_tolerance else "**NO**"
        lines.append(
            f"| {r.sheet} | {r.row} | {r.label} | "
            f"{r.filed:,.2f} | {r.generated:,.2f} | "
            f"{r.delta:+,.2f} | {ok} |"
        )
    total = len(results)
    matched = sum(1 for r in results if r.within_tolerance)
    pct = matched / total * 100 if total else 0
    lines.append(f"\n**{matched}/{total} cells within tolerance ({pct:.1f}%)**")
    return "\n".join(lines)
