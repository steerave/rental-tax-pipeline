"""Bootstrap vendor mappings from Rent QC owner statements.

For each Rent QC transaction, the category is embedded in the description
(e.g., "102 - Rent Income - January 2024"). This module:
1. Collects all unique vendor -> category assignments across Rent QC reports.
2. Records them in the vendor_mapping format with high confidence.
3. Optionally reconciles summed categories against a filed XLSX.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import openpyxl

from taxauto.aggregate.rentqc_to_template import map_rentqc_category
from taxauto.categorize.mapper import normalize_vendor
from taxauto.parsers.rent_qc import RentQCReport, RentQCTransaction


@dataclass
class BootstrapReport:
    total_transactions: int = 0
    categories_seen: int = 0
    vendors_learned: int = 0
    reconciliation: Dict[str, Dict[str, Dict[str, Decimal]]] = field(default_factory=dict)
    # reconciliation[property][category] = {"parsed": X, "filed": Y, "delta": Z}


def _collect_rentqc_transactions(reports: Iterable[RentQCReport]) -> List[RentQCTransaction]:
    """Flatten all transactions from all properties from all reports."""
    all_txns: List[RentQCTransaction] = []
    for report in reports:
        for prop in report.properties:
            all_txns.extend(prop.transactions)
    return all_txns


def learn_vendors_from_rentqc(
    reports: Iterable[RentQCReport],
    mapping: Dict[str, Any],
    *,
    year: int,
    confidence: float = 0.95,
) -> BootstrapReport:
    """Learn vendor -> category mappings from Rent QC transactions.

    Updates ``mapping`` in place and returns a report.
    """
    all_txns = _collect_rentqc_transactions(reports)
    report = BootstrapReport(total_transactions=len(all_txns))

    vendors = mapping.setdefault("vendors", {})
    categories_seen: set = set()

    for txn in all_txns:
        if not txn.category:
            continue
        categories_seen.add(txn.category)

        # Use the payee as the vendor key for vendor_mapping learning.
        key = normalize_vendor(txn.payee) if txn.payee else None
        if not key:
            continue

        entry = vendors.get(key)
        if entry is None:
            vendors[key] = {
                "category": txn.category,
                "confidence": confidence,
                "occurrences": 1,
                "learned_from": [year],
                "ambiguous": False,
                "source": "rentqc_bootstrap",
            }
            report.vendors_learned += 1
        else:
            entry["occurrences"] = int(entry.get("occurrences", 0)) + 1
            learned_from = entry.setdefault("learned_from", [])
            if year not in learned_from:
                learned_from.append(year)
            # If same vendor seen with different category, mark ambiguous
            if entry.get("category") and entry["category"] != txn.category:
                entry["ambiguous"] = True

    report.categories_seen = len(categories_seen)
    return report


def reconcile_against_filed_xlsx(
    reports: Iterable[RentQCReport],
    filed_xlsx_path: Path,
    rentqc_mapping: Dict[str, Optional[str]],
    *,
    tax_year: int,
) -> Dict[str, Dict[str, Dict[str, Decimal]]]:
    """Sum Rent QC transactions by (property, template_category) and compare
    against the filed XLSX cell values.

    Returns {property: {template_category: {"parsed": X, "filed": Y, "delta": Z}}}
    """
    # Sum from Rent QC
    parsed_totals: Dict[str, Dict[str, Decimal]] = {}
    for report in reports:
        for prop in report.properties:
            for txn in prop.transactions:
                if txn.date.year != tax_year:
                    continue
                if not txn.category:
                    continue
                template_cat = map_rentqc_category(txn.category, rentqc_mapping)
                if template_cat is None:
                    continue
                amount = txn.cash_in if txn.cash_in else (txn.cash_out if txn.cash_out else Decimal("0"))
                parsed_totals.setdefault(prop.name, {}).setdefault(template_cat, Decimal("0"))
                parsed_totals[prop.name][template_cat] += amount

    # Load filed XLSX for comparison
    result: Dict[str, Dict[str, Dict[str, Decimal]]] = {}
    if not filed_xlsx_path.exists():
        # No filed XLSX to compare -- just return parsed totals
        for prop_name, cats in parsed_totals.items():
            result[prop_name] = {}
            for cat, parsed_val in cats.items():
                result[prop_name][cat] = {
                    "parsed": parsed_val,
                    "filed": Decimal("0"),
                    "delta": parsed_val,
                }
        return result

    wb = openpyxl.load_workbook(filed_xlsx_path, data_only=True)
    from taxauto.writers.ltr_writer import scan_label_rows

    for prop_name, cats in parsed_totals.items():
        result[prop_name] = {}
        if prop_name not in wb.sheetnames:
            for cat, parsed_val in cats.items():
                result[prop_name][cat] = {
                    "parsed": parsed_val,
                    "filed": Decimal("0"),
                    "delta": parsed_val,
                }
            continue

        ws = wb[prop_name]
        label_rows = scan_label_rows(ws)

        for cat, parsed_val in cats.items():
            filed_val = Decimal("0")
            # Find the row for this category
            target_lower = cat.lower()
            for label, row in label_rows.items():
                if label.lower() == target_lower:
                    cell_val = ws[f"D{row}"].value
                    if cell_val is not None:
                        try:
                            filed_val = Decimal(str(cell_val))
                        except Exception:
                            pass
                    break

            result[prop_name][cat] = {
                "parsed": parsed_val,
                "filed": filed_val,
                "delta": parsed_val - filed_val,
            }

    return result
