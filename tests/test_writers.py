"""Tests for the Phase 2 STR, LTR, and transactions-tab Excel writers."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from taxauto.writers.str_writer import write_str_workbook, ALL_STR_ROWS
from taxauto.writers.ltr_writer import write_ltr_workbook, scan_label_rows
from taxauto.writers.transactions_tab import append_transactions_tab


STR_TEMPLATE = Path(__file__).parent / "fixtures" / "output_templates" / "str_template.xlsx"
LTR_TEMPLATE = Path(__file__).parent / "fixtures" / "output_templates" / "ltr_template.xlsx"


# ── STR writer tests ─────────────────────────────────────────────────────


def test_str_writes_per_property_totals(tmp_path: Path) -> None:
    out = tmp_path / "str_2025.xlsx"
    totals = {
        "15 Belden": {
            "Sales revenue": Decimal("57739.81"),
            "Utilities": Decimal("5253.58"),
        },
        "27 Farmstead Dr": {"Sales revenue": Decimal("30000.00")},
    }
    write_str_workbook(
        template_path=STR_TEMPLATE,
        output_path=out,
        per_property_totals=totals,
        year=2025,
    )

    wb = openpyxl.load_workbook(out)
    assert wb["15 Belden"]["D5"].value == 57739.81
    assert wb["15 Belden"]["D31"].value == 5253.58
    assert wb["27 Farmstead Dr"]["D5"].value == 30000.00
    assert wb["15 Belden"]["D4"].value == 2025


def test_str_preserves_formulas(tmp_path: Path) -> None:
    out = tmp_path / "str_2025.xlsx"
    write_str_workbook(
        template_path=STR_TEMPLATE,
        output_path=out,
        per_property_totals={},
        year=2025,
    )

    wb = openpyxl.load_workbook(out)
    ws = wb["15 Belden"]
    # Total Revenues at D10 should be a formula
    d10 = ws["D10"].value
    assert d10 is not None and str(d10).startswith("="), f"D10 should be a formula, got {d10}"


def test_str_forces_other_label(tmp_path: Path) -> None:
    out = tmp_path / "str_2025.xlsx"
    write_str_workbook(
        template_path=STR_TEMPLATE,
        output_path=out,
        per_property_totals={},
        year=2025,
    )

    wb = openpyxl.load_workbook(out)
    assert wb["15 Belden"]["C33"].value == "other"


def test_str_skips_zero_values(tmp_path: Path) -> None:
    out = tmp_path / "str_2025.xlsx"
    totals = {
        "15 Belden": {
            "Sales revenue": Decimal("1000.00"),
            "Advertising": Decimal("0"),
        },
    }
    write_str_workbook(
        template_path=STR_TEMPLATE,
        output_path=out,
        per_property_totals=totals,
        year=2025,
    )
    wb = openpyxl.load_workbook(out)
    assert wb["15 Belden"]["D5"].value == 1000.00
    assert wb["15 Belden"]["D13"].value is None  # zero should not be written


def test_str_sets_property_name_in_b1(tmp_path: Path) -> None:
    out = tmp_path / "str_2025.xlsx"
    write_str_workbook(
        template_path=STR_TEMPLATE,
        output_path=out,
        per_property_totals={},
        year=2025,
    )
    wb = openpyxl.load_workbook(out)
    assert wb["15 Belden"]["B1"].value == "15 Belden"
    assert wb["27 Farmstead Dr"]["B1"].value == "27 Farmstead Dr"


# ── LTR writer tests ─────────────────────────────────────────────────────


def test_ltr_writes_with_dynamic_row_scan(tmp_path: Path) -> None:
    out = tmp_path / "ltr_2025.xlsx"
    totals = {
        "1015 39th St": {
            "Sales revenue": Decimal("96533.92"),
            "Management Fees": Decimal("9653.40"),
            "Lawn and Snow Care": Decimal("1950.00"),
        },
        "1210 College Ave": {
            "Sales revenue": Decimal("50000.00"),
            "Lawn and Snow Care": Decimal("800.00"),  # lowercase 'c' in template
        },
    }
    write_ltr_workbook(
        template_path=LTR_TEMPLATE,
        output_path=out,
        per_property_totals=totals,
        year=2025,
    )

    wb = openpyxl.load_workbook(out)

    # Verify 1015 39th St
    ws1 = wb["1015 39th St"]
    label_rows_1015 = scan_label_rows(ws1)
    mgmt_row = next(r for l, r in label_rows_1015.items() if l.lower() == "management fees")
    assert ws1[f"D{mgmt_row}"].value == 9653.40

    # Verify 1210 College Ave — "Lawn and Snow care" (lowercase c) should still match
    ws2 = wb["1210 College Ave"]
    label_rows_1210 = scan_label_rows(ws2)
    lawn_row = next((r for l, r in label_rows_1210.items() if "lawn" in l.lower()), None)
    assert lawn_row is not None, "Lawn and Snow Care/care not found on 1210"
    assert ws2[f"D{lawn_row}"].value == 800.00

    # Year updated
    assert ws1["D4"].value == 2025


def test_ltr_preserves_formulas(tmp_path: Path) -> None:
    out = tmp_path / "ltr_2025.xlsx"
    write_ltr_workbook(
        template_path=LTR_TEMPLATE,
        output_path=out,
        per_property_totals={},
        year=2025,
    )
    wb = openpyxl.load_workbook(out)
    ws = wb["1015 39th St"]
    d10 = ws["D10"].value
    assert d10 is not None and str(d10).startswith("="), f"D10 should be a formula, got {d10}"


def test_ltr_warns_on_missing_category(tmp_path: Path) -> None:
    out = tmp_path / "ltr_2025.xlsx"
    totals = {
        "308 Lincoln Ave": {
            "Nonexistent Category XYZ": Decimal("100.00"),
        },
    }
    with pytest.warns(UserWarning, match="not found on sheet"):
        write_ltr_workbook(
            template_path=LTR_TEMPLATE,
            output_path=out,
            per_property_totals=totals,
            year=2025,
        )


def test_ltr_sets_property_name_in_b1(tmp_path: Path) -> None:
    out = tmp_path / "ltr_2025.xlsx"
    write_ltr_workbook(
        template_path=LTR_TEMPLATE,
        output_path=out,
        per_property_totals={},
        year=2025,
    )
    wb = openpyxl.load_workbook(out)
    assert wb["1015 39th St"]["B1"].value == "1015 39th St"


# ── Transactions tab tests ───────────────────────────────────────────────


def test_transactions_tab_appended(tmp_path: Path) -> None:
    out = tmp_path / "workbook.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "15 Belden"
    wb.save(out)

    txns = [
        {
            "date": date(2024, 1, 5),
            "source": "checking",
            "description": "HOME DEPOT",
            "amount": Decimal("-150.00"),
            "template_category": "Supplies",
            "notes": "",
        },
        {
            "date": date(2024, 2, 10),
            "source": "credit 1091",
            "description": "COSTCO",
            "amount": Decimal("-230.00"),
            "template_category": "Supplies",
            "notes": "",
        },
    ]
    append_transactions_tab(out, sheet_name="15 Belden", transactions=txns)

    wb2 = openpyxl.load_workbook(out)
    assert "15 Belden_txns" in wb2.sheetnames
    ws = wb2["15 Belden_txns"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("date", "source", "description", "amount", "template_category", "notes")
    assert len(rows) == 3  # header + 2 data rows


def test_transactions_tab_replaces_existing(tmp_path: Path) -> None:
    out = tmp_path / "workbook.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "15 Belden"
    wb.save(out)

    txns1 = [{"date": date(2024, 1, 1), "amount": Decimal("100")}]
    append_transactions_tab(out, sheet_name="15 Belden", transactions=txns1)

    txns2 = [{"date": date(2024, 2, 1), "amount": Decimal("200")}]
    append_transactions_tab(out, sheet_name="15 Belden", transactions=txns2)

    wb2 = openpyxl.load_workbook(out)
    assert wb2.sheetnames.count("15 Belden_txns") == 1
    ws = wb2["15 Belden_txns"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 2  # header + 1 replacement row
