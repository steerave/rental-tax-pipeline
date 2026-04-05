from pathlib import Path

import openpyxl

from taxauto.writers.template_copy import copy_and_clear_values


def test_copies_template_and_clears_value_cells(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B1"] = "15 Belden"
    ws["B4"] = "Revenue"
    ws["D4"] = 2024
    ws["C5"] = "Sales revenue"
    ws["D5"] = 57739.81
    ws["C13"] = "Advertising"
    ws["D13"] = 222.57
    ws["D10"] = "=SUM(D5:D9)"  # formula
    wb.save(src)

    dst = tmp_path / "dst.xlsx"
    copy_and_clear_values(src, dst)

    wb2 = openpyxl.load_workbook(dst)
    ws2 = wb2.active
    assert ws2["B1"].value == "15 Belden"       # label preserved
    assert ws2["C5"].value == "Sales revenue"   # label preserved
    assert ws2["D5"].value is None              # value cleared
    assert ws2["D13"].value is None             # value cleared
    assert ws2["D4"].value is None              # year cleared (writer sets it later)
    assert ws2["D10"].value == "=SUM(D5:D9)"    # formula preserved
