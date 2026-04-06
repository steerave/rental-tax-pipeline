"""Tests for the filed-vs-generated comparison."""

from pathlib import Path

import openpyxl
import pytest

from taxauto.verify.compare import compare_workbooks, ComparisonResult


def _make_workbook(path: Path, sheets: dict) -> None:
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, cells in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for cell, value in cells.items():
            ws[cell] = value
    wb.save(path)


def test_compare_detects_matching_cells(tmp_path: Path) -> None:
    filed = tmp_path / "filed.xlsx"
    generated = tmp_path / "generated.xlsx"
    _make_workbook(filed, {"Sheet1": {"C5": "Sales revenue", "D5": 1000.00}})
    _make_workbook(generated, {"Sheet1": {"C5": "Sales revenue", "D5": 1000.50}})

    results = compare_workbooks(filed, generated)
    assert len(results) == 1
    assert results[0].within_tolerance is True


def test_compare_detects_mismatch(tmp_path: Path) -> None:
    filed = tmp_path / "filed.xlsx"
    generated = tmp_path / "generated.xlsx"
    _make_workbook(filed, {"Sheet1": {"C13": "Utilities", "D13": 5000.00}})
    _make_workbook(generated, {"Sheet1": {"C13": "Utilities", "D13": 3000.00}})

    results = compare_workbooks(filed, generated)
    assert len(results) == 1
    assert results[0].within_tolerance is False


def test_compare_handles_missing_sheet(tmp_path: Path) -> None:
    filed = tmp_path / "filed.xlsx"
    generated = tmp_path / "generated.xlsx"
    _make_workbook(filed, {"Sheet1": {"D5": 100}, "Sheet2": {"D5": 200}})
    _make_workbook(generated, {"Sheet1": {"D5": 100}})

    results = compare_workbooks(filed, generated)
    sheet2_results = [r for r in results if r.sheet == "Sheet2"]
    assert len(sheet2_results) >= 1
    assert sheet2_results[0].generated == 0.0
