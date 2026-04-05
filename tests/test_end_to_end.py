"""End-to-end pipeline test against a synthetic fixture year.

Exercises: extract → categorize → (hand-written decisions) → build
against real text PDFs generated at runtime with reportlab, and a real
blank Excel template.

Sheets push/pull are covered separately in test_sheets.py and are not
part of this deterministic path.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pytest
import yaml

from taxauto.cli import (
    _cache_path,
    _write_json,
    cmd_build,
    cmd_categorize,
    cmd_extract,
)
from taxauto.config import load_config


BANK_STATEMENT_TEXT = (
    "ACME BANK - STATEMENT\n"
    "Account: ****9999\n"
    "Period: 01/01/2025 - 01/31/2025\n"
    "Beginning Balance: $5,000.00\n"
    "Ending Balance: $4,650.00\n"
    "Total Deposits: $500.00\n"
    "Total Withdrawals: $850.00\n"
    "\n"
    "Date   Description                                  Amount    Balance\n"
    "01/05  POS PURCHASE HOME DEPOT #001                 -150.00  4,850.00\n"
    "01/10  ACH DEBIT AIRBNB PAYOUT FEE                   -50.00  4,800.00\n"
    "01/15  DEPOSIT RENT 123 MAIN ST                     500.00  5,300.00\n"
    "01/20  POS PURCHASE SAFEWAY                         -100.00  5,200.00\n"
    "01/28  ACH DEBIT FURNITURE STORE                    -550.00  4,650.00\n"
)


PM_REPORT_TEXT = (
    "ACME PROPERTY MANAGEMENT\n"
    "Owner Statement - 123 Main St\n"
    "Period: 01/01/2025 - 12/31/2025\n"
    "\n"
    "Date        Description                          Category            Amount\n"
    "01/15/2025  Rent Received - Tenant A              Rental Income        500.00\n"
    "01/20/2025  Landscaping                           Maintenance         -120.00\n"
    "\n"
    "Total Income: 500.00\n"
    "Total Expenses: -120.00\n"
    "Net to Owner: 380.00\n"
)


def _make_text_pdf(path: Path, text: str) -> None:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=letter)
    c.setFont("Courier", 9)
    text_obj = c.beginText(36, 750)
    for line in text.splitlines():
        text_obj.textLine(line)
    c.drawText(text_obj)
    c.showPage()
    c.save()


def _make_blank_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(["Date", "Description", "Amount", "Category"])
    wb.save(path)


@pytest.fixture()
def project(tmp_path: Path) -> Path:
    """Build a minimal project tree under tmp_path."""
    root = tmp_path / "proj"
    root.mkdir()

    # Config
    (root / "config.yaml").write_text(
        """
templates:
  str: templates/str.xlsx
  ltr: templates/ltr.xlsx
year_paths:
  root: years/{year}
  bank_inputs: years/{year}/inputs/bank
  pm_ltr_inputs: years/{year}/inputs/pm_ltr
  pm_str_inputs: years/{year}/inputs/pm_str
  intermediate: years/{year}/intermediate
  outputs: years/{year}/outputs
  review_log: years/{year}/review_log.csv
categories:
  primary: [STR, LTR, Personal, Skip, Split]
  str_subcategories: []
  ltr_subcategories: []
property_managers:
  ltr: {name_contains: [ACME PROPERTY]}
  str: {name_contains: []}
bootstrap:
  date_window_days: 3
  description_fuzzy_threshold: 0.72
  min_confidence_to_auto_learn: 0.80
""",
        encoding="utf-8",
    )

    # Empty vendor mapping — pipeline must push everything to review.
    (root / "vendor_mapping.yaml").write_text("vendors: {}\n", encoding="utf-8")

    # Templates
    _make_blank_template(root / "templates" / "str.xlsx")
    _make_blank_template(root / "templates" / "ltr.xlsx")

    # Synthetic PDFs
    _make_text_pdf(root / "years" / "2025" / "inputs" / "bank" / "acme_jan.pdf", BANK_STATEMENT_TEXT)
    _make_text_pdf(root / "years" / "2025" / "inputs" / "pm_ltr" / "acme_pm.pdf", PM_REPORT_TEXT)

    return root


def test_end_to_end_extract_categorize_build(project: Path) -> None:
    cfg = load_config(project / "config.yaml")
    year = 2025

    # --- extract ---
    rc = cmd_extract(cfg, year)
    assert rc == 0
    paths = cfg.paths_for_year(year)
    assert (_cache_path(paths, "bank_transactions.json")).exists()
    assert (_cache_path(paths, "pm_ltr_entries.json")).exists()

    # --- categorize (empty mapping → everything goes to review) ---
    rc = cmd_categorize(cfg, year)
    assert rc == 0
    import json
    queue = json.loads((_cache_path(paths, "review_queue.json")).read_text(encoding="utf-8"))
    auto = json.loads((_cache_path(paths, "auto_tagged.json")).read_text(encoding="utf-8"))
    assert len(auto) == 0
    assert len(queue) == 5  # all 5 bank transactions routed to review

    # --- hand-write review decisions (simulate bookkeeper tagging) ---
    decisions = []
    for q in queue:
        desc = q["transaction"]["description"]
        if "HOME DEPOT" in desc or "AIRBNB" in desc:
            category = "STR"
        elif "FURNITURE" in desc or "RENT 123 MAIN" in desc:
            category = "LTR"
        else:
            category = "Personal"
        decisions.append(
            {
                "row_id": f"2025-{desc[:5]}",
                "date": q["transaction"]["date"],
                "description": desc,
                "amount": q["transaction"]["amount"],
                "category": category,
                "year": year,
            }
        )
    _write_json(_cache_path(paths, "review_decisions.json"), decisions)

    # --- build ---
    rc = cmd_build(cfg, year)
    assert rc == 0
    str_out = paths.outputs / "str_2025.xlsx"
    ltr_out = paths.outputs / "ltr_2025.xlsx"
    assert str_out.exists()
    assert ltr_out.exists()

    # STR workbook should contain HOME DEPOT + AIRBNB (2 rows after header).
    wb = openpyxl.load_workbook(str_out)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    descs = [r[1] for r in rows]
    assert any("HOME DEPOT" in d for d in descs)
    assert any("AIRBNB" in d for d in descs)
    assert len(rows) == 2

    # LTR workbook: 2 PM entries + FURNITURE STORE (bank).
    # The bank "RENT 123 MAIN ST" deposit collides with the PM "Rent Received"
    # line (same $500 amount, same 01/15 date), so the double-count guard
    # MUST drop it. This is exactly the regression we want the e2e to catch.
    wb = openpyxl.load_workbook(ltr_out)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    descs = [r[1] for r in rows]
    assert "Rent Received - Tenant A" in descs
    assert "Landscaping" in descs
    assert any("FURNITURE" in d for d in descs)
    # Double-count guard must have removed the bank rent deposit.
    assert not any("RENT 123 MAIN" in d for d in descs)
    assert len(rows) == 3
